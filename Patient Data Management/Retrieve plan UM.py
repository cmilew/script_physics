from connect import *
import openpyxl

# Entry and exit excel files
excel_file_path = '\\\\nas-01\\PHY_PROC\\Radiotherapie\\TPS\\Raystation\\Scripting\\Scripts\\patients_to_retrieve_plan_UM_from.xlsx'
result_excel_file_path = '\\\\nas-01\\PHY_PROC\\Radiotherapie\\TPS\\Raystation\\Scripting\\Scripts\\UM_results.xlsx'

# Open patient in RayStation
patient_db = get_current("PatientDB")

# Get excel file with patients ID and dosimetric index to retrieve
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Get excel file for the results
wb_results = openpyxl.load_workbook(result_excel_file_path)
results_sheet = wb_results.active

# Get patients ID to retrieve
patient_row = 2
result_row = 2
stop_loop = False
while (sheet.cell(row=patient_row, column=1)).value:
    # Get ID from entry excel file and write it in first colum of results excel file
    patient_ID = (sheet.cell(row=patient_row, column=1)).value
    print(patient_ID)
    # Delete potential space in front or at the end of patient's ID
    patient_ID = patient_ID.strip()
    active_cell_results = results_sheet.cell(row=result_row, column=1)
    active_cell_results.value = patient_ID
    active_cell_results.alignment = openpyxl.styles.alignment.Alignment(horizontal='center',
                                                                        vertical='center')
    # ID check
    if len(str(patient_ID)) != 11:
        issue = 'Le format de l\'ID du patient ' + str(patient_ID) + ' n\'est pas correct.'
        print(issue)
        stop_loop = True
    else:
        patient_ID_figures = patient_ID[0:9]
        patient_ID_letters = patient_ID[9:11]
        try:
            patient_ID_figures = int(patient_ID_figures)
        except ValueError:
            issue = 'Le format de l\'ID du patient ' + str(patient_ID) + ' n\'est pas correct.'
            stop_loop = True
        if not patient_ID_letters.isalpha():
            issue = 'Le format de l\'ID du patient ' + str(patient_ID) + ' n\'est pas correct.'
            stop_loop = True
    # Query for patient ID in both primary and secondary database
    if not stop_loop:
        try:
            info = patient_db.QueryPatientInfo(Filter={'PatientID': patient_ID}, UseIndexService=True)
            patient = patient_db.LoadPatient(PatientInfo=info[0], AllowPatientUpgrade=True)
        except IndexError:
            issue = 'Patient avec ID ' + str(patient_ID) + ' pas trouvé dans RayStation.'
            print(issue)
            stop_loop = True

        if not stop_loop:
            # Case check
            # If only one case, script opens it automatically, if 0 or more cases found for this patient, an issue is
            # generated and script continues with next patient on list
            nb_of_case = len(patient.Cases)
            if nb_of_case == 0:
                issue = 'Pas de case trouvé pour patient ' + str(patient_ID)
                print(issue)
                stop_loop = True
            elif nb_of_case > 1:
                issue = 'Plusieurs cases trouvés pour patient ' + str(patient_ID)
                print(issue)
                stop_loop = True
            else:
                patient.Cases[0].SetCurrent()
                case = get_current('Case')

            # Plan check
            # Script only opens plan that are approved
            # If no plan approved or more than one plan approved for a patient, an issue is generated for this patient
            # and script continues with next patient on list
            # For a plan to be approved, every beamset in it has must be approved
            if not stop_loop:
                list_of_approved_plans = []
                if len(case.TreatmentPlans) > 0:
                    for treatment_plan in case.TreatmentPlans:
                        nb_of_beam_set = len(treatment_plan.BeamSets)
                        # if no beamset found for this plan, goes to next plan
                        if nb_of_beam_set == 0:
                            pass
                        else:
                            for beam_set in treatment_plan.BeamSets:
                                if beam_set.Review is not None:
                                    plan_is_approved = True
                                else:
                                    plan_is_approved = False
                            if plan_is_approved:
                                list_of_approved_plans.append(treatment_plan.Name)
                    if len(list_of_approved_plans) == 0:
                        issue = 'Pas de plan approuvé trouvé pour patient ' + str(patient_ID)
                        print(issue)
                        stop_loop = True
                    elif len(list_of_approved_plans) > 1:
                        issue = 'Plusieurs plans approuvés trouvés pour patient ' + str(patient_ID)
                        print(issue)
                        stop_loop = True
                    else:
                        approved_plan = case.TreatmentPlans[list_of_approved_plans[0]]
                else:
                    issue = 'Pas de plan trouvé pour patient ' + str(patient_ID)
                    print(issue)
                    stop_loop = True

                # If no issue encountered with plan and case, script retrieves UM of each for each beamset
                if not stop_loop:
                    # Enters plan name in 2nd column of results excel file
                    active_cell_results = results_sheet.cell(row=result_row, column=2)
                    active_cell_results.value = approved_plan.Name
                    active_cell_results.alignment = openpyxl.styles.alignment.Alignment(horizontal='center',
                                                                                        vertical='center')
                    for beam_set in approved_plan.BeamSets:
                        # Enters beam set name in 3rd column of excel file
                        active_cell_results = results_sheet.cell(row=result_row, column=3)
                        active_cell_results.value = beam_set.DicomPlanLabel
                        active_cell_results.alignment = openpyxl.styles.alignment.Alignment(horizontal='center',
                                                                                            vertical='center')
                        if len(beam_set.Beams) ==0:
                            text = 'Pas de faisceau trouvé pour ce beam set.'
                            active_cell_results = results_sheet.cell(row=result_row, column=3)
                            active_cell_results.value = text
                        else:
                            for beam in beam_set.Beams:
                                active_cell_results = results_sheet.cell(row=result_row, column=4)
                                active_cell_results.value = beam.Name
                                active_cell_results.alignment = openpyxl.styles.alignment.Alignment(horizontal='center',
                                                                                       vertical='center')
                                active_cell_results = results_sheet.cell(row=result_row, column=5)
                                active_cell_results.value = beam.Description
                                active_cell_results.alignment = openpyxl.styles.alignment.Alignment(horizontal='center',
                                                                                                    vertical='center')
                                active_cell_results = results_sheet.cell(row=result_row, column=6)
                                active_cell_results.value = beam.BeamMU
                                active_cell_results.alignment = openpyxl.styles.alignment.Alignment(horizontal='center',
                                                                                                    vertical='center')
                                result_row += 1

    if stop_loop:
        issue_cell = results_sheet.cell(row=result_row, column=2)
        issue_cell.value = issue
        result_row += 1
    stop_loop = False
    patient_row +=1
    wb_results.save(result_excel_file_path)

wb_results.save(result_excel_file_path)

